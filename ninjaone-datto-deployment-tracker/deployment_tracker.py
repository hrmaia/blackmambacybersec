"""
deployment_tracker.py
----------------------
Cross-references a NinjaOne RMM device export against a legacy DattoRMM device
export, per client organization, to track migration progress.

Built while migrating ~22 MSP client organizations from DattoRMM to NinjaOne.
Naming conventions for the same physical device often differed between the two
platforms (site-code prefixes, asset tags, trailing suffixes), so matching is
config-driven and supports multiple rule types per organization, applied in
priority order until a match is found or all rules are exhausted.

Usage:
    python deployment_tracker.py --config config.json \
        --ninja ninja_export.csv --datto datto_export.csv \
        --out report.xlsx

Expected CSV columns (case-insensitive, flexible order):
    Organization, Hostname, Last Seen  (+ any extra columns are ignored)

NOTE: This is a reconstructed / sanitized version for portfolio purposes.
Organization names, domains, and file paths are placeholders.
"""

import argparse
import json
import re
from dataclasses import dataclass, field
from datetime import datetime, timezone
from difflib import SequenceMatcher
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Status constants (drive both logic and Excel conditional coloring)
# ---------------------------------------------------------------------------
STATUS_MIGRATED = "Migrated"          # present in both platforms
STATUS_NINJA_ONLY = "NinjaOne Only"   # deployed, Datto not yet decommissioned
STATUS_DATTO_ONLY = "Datto Only"      # not yet migrated
STATUS_REVIEW = "Needs Review"        # fuzzy match below confidence threshold

STATUS_FILL = {
    STATUS_MIGRATED: "C6EFCE",     # green
    STATUS_NINJA_ONLY: "FFEB9C",   # yellow
    STATUS_DATTO_ONLY: "FFC7CE",   # red
    STATUS_REVIEW: "FFD966",       # amber
}


@dataclass
class MatchRule:
    type: str
    priority: int
    suffix: str = ""
    prefix: str = ""
    threshold: float = 0.85


@dataclass
class OrgConfig:
    name: str
    domain: str
    match_rules: list = field(default_factory=list)


def load_config(path: str) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        raw = json.load(f)

    orgs = {}
    for o in raw["organizations"]:
        rules = [MatchRule(**r) for r in o.get("match_rules", [])]
        rules.sort(key=lambda r: r.priority)
        orgs[o["name"]] = OrgConfig(name=o["name"], domain=o["domain"], match_rules=rules)
    return {"organizations": orgs, "output": raw.get("output", {})}


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Map flexible source-export column names onto a fixed internal schema."""
    colmap = {}
    for col in df.columns:
        c = col.strip().lower()
        if c in ("organization", "org", "client", "site"):
            colmap[col] = "Organization"
        elif c in ("hostname", "device name", "device", "computer name"):
            colmap[col] = "Hostname"
        elif c in ("last seen", "lastseen", "last check-in", "last checkin"):
            colmap[col] = "LastSeen"
    df = df.rename(columns=colmap)
    for required in ("Organization", "Hostname"):
        if required not in df.columns:
            raise ValueError(f"Required column '{required}' not found in export.")
    if "LastSeen" not in df.columns:
        df["LastSeen"] = None
    return df[["Organization", "Hostname", "LastSeen"]].copy()


def apply_rule(hostname: str, rule: MatchRule) -> str:
    """Transform a hostname according to a single match rule, for comparison."""
    h = hostname.strip().upper()
    if rule.type == "exact":
        return h
    if rule.type == "strip_suffix" and rule.suffix:
        suf = rule.suffix.upper()
        return h[: -len(suf)] if h.endswith(suf) else h
    if rule.type == "prefix" and rule.prefix:
        pre = rule.prefix.upper()
        return h[len(pre):] if h.startswith(pre) else h
    if rule.type == "fuzzy":
        return h  # fuzzy comparison handled separately in match_hostname
    return h


def fuzzy_ratio(a: str, b: str) -> float:
    return SequenceMatcher(None, a.upper(), b.upper()).ratio()


def match_hostname(hostname: str, candidates: list, rules: list):
    """
    Try each configured rule in priority order. Returns (matched_hostname, status)
    or (None, None) if nothing matches even fuzzily.
    """
    for rule in rules:
        if rule.type == "fuzzy":
            best, best_score = None, 0.0
            for c in candidates:
                score = fuzzy_ratio(hostname, c)
                if score > best_score:
                    best, best_score = c, score
            if best and best_score >= rule.threshold:
                status = STATUS_MIGRATED if best_score > 0.97 else STATUS_REVIEW
                return best, status
        else:
            transformed = apply_rule(hostname, rule)
            for c in candidates:
                if apply_rule(c, rule) == transformed:
                    return c, STATUS_MIGRATED
    return None, None


def build_org_report(org_name: str, ninja_df: pd.DataFrame, datto_df: pd.DataFrame, rules: list) -> pd.DataFrame:
    ninja_hosts = ninja_df.loc[ninja_df["Organization"] == org_name, "Hostname"].tolist()
    datto_hosts = datto_df.loc[datto_df["Organization"] == org_name, "Hostname"].tolist()

    rows = []
    matched_datto = set()

    for nh in ninja_hosts:
        match, status = match_hostname(nh, datto_hosts, rules)
        if match:
            matched_datto.add(match)
            rows.append({"Hostname (NinjaOne)": nh, "Hostname (Datto)": match, "Status": status})
        else:
            rows.append({"Hostname (NinjaOne)": nh, "Hostname (Datto)": "", "Status": STATUS_NINJA_ONLY})

    for dh in datto_hosts:
        if dh not in matched_datto:
            rows.append({"Hostname (NinjaOne)": "", "Hostname (Datto)": dh, "Status": STATUS_DATTO_ONLY})

    return pd.DataFrame(rows)


def write_excel(reports: dict, out_path: str):
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="203864")

    summary_ws.append(["Organization", "Migrated", "NinjaOne Only", "Datto Only", "Needs Review", "Total Devices"])
    for cell in summary_ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for org_name, df in reports.items():
        counts = df["Status"].value_counts().to_dict()
        summary_ws.append([
            org_name,
            counts.get(STATUS_MIGRATED, 0),
            counts.get(STATUS_NINJA_ONLY, 0),
            counts.get(STATUS_DATTO_ONLY, 0),
            counts.get(STATUS_REVIEW, 0),
            len(df),
        ])

        ws = wb.create_sheet(title=org_name[:31])  # Excel sheet name limit
        ws.append(list(df.columns))
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        for _, row in df.iterrows():
            ws.append(list(row))
            status_cell = ws.cell(row=ws.max_row, column=3)
            fill_color = STATUS_FILL.get(row["Status"])
            if fill_color:
                status_cell.fill = PatternFill("solid", fgColor=fill_color)

        for col_idx in range(1, len(df.columns) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 28

    for col_idx in range(1, 7):
        summary_ws.column_dimensions[get_column_letter(col_idx)].width = 20

    wb.save(out_path)


def update_history(history_path: str, reports: dict, max_snapshots: int):
    """Append today's per-org counts to a rolling JSON history file."""
    history_file = Path(history_path)
    history = json.loads(history_file.read_text()) if history_file.exists() else []

    snapshot = {
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "orgs": {
            org: df["Status"].value_counts().to_dict()
            for org, df in reports.items()
        },
    }
    history.append(snapshot)
    history = history[-max_snapshots:]
    history_file.write_text(json.dumps(history, indent=2))


def main():
    parser = argparse.ArgumentParser(description="NinjaOne vs DattoRMM deployment tracker")
    parser.add_argument("--config", required=True, help="Path to config.json")
    parser.add_argument("--ninja", required=True, help="Path to NinjaOne device export CSV")
    parser.add_argument("--datto", required=True, help="Path to DattoRMM device export CSV")
    parser.add_argument("--out", default="report.xlsx", help="Output Excel report path")
    args = parser.parse_args()

    cfg = load_config(args.config)
    ninja_df = normalize_columns(pd.read_csv(args.ninja))
    datto_df = normalize_columns(pd.read_csv(args.datto))

    reports = {}
    for org_name, org_cfg in cfg["organizations"].items():
        reports[org_name] = build_org_report(org_name, ninja_df, datto_df, org_cfg.match_rules)

    write_excel(reports, args.out)

    history_path = cfg["output"].get("history_file", "history.json")
    max_snapshots = cfg["output"].get("max_snapshots_kept", 12)
    update_history(history_path, reports, max_snapshots)

    print(f"Report written to {args.out}")
    print(f"History updated at {history_path}")


if __name__ == "__main__":
    main()
